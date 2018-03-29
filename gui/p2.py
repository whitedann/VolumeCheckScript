import sys
import requests
import os
import sspyrs
import openpyxl
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
import getpass

class Builder:

    def __init__(self,barcode,customer):
        self.barcode = barcode
        self.customer = customer
        
        
    #collects query info from user and changes working directory 
    def init(self):
    
        #self.barcode = input('Enter BARCODE: ')
        #self.customer = input('Enter Customer (optional): ')
        
        #Work directory
        os.chdir('W:\\Employees\Danny\dev')

        #Home directory
        #os.chdir('C:\\Users\Dan\Desktop\idtdocs')
    
        return 0;

    #collects login info and starts session timer
    def login(self):

        self.user = input('type user (don\'t scan badge): ')
        self.password = getpass.getpass('type password (hidden): ')
    
        return 0;

    #get template file from idtdna.mastercontrol.com using login info
    def getTemplate(self):

        #the four parameters required to log into master control...the middle two might not be needed...
        payload = {'username':self.user,'username2':'','initialRequest':'','password':self.password}

        with requests.Session() as s:
            p = s.post('https://idtdna.mastercontrol.com/mc/login/index.cfm?action=login',data=payload)

            #*****NOTE:If you keep getting Error 1 even when with good login credentials, then then the following URL may be out of date!!#
            r = s.get('https://idtdna.mastercontrol.com/mc/Main/MASTERControl/Organizer/view_file.cfm?id=XBCYSHLULRA4XBIBB7')

            #Checks to see if the downloaded file is correct by assessing its relative size.
            if len(r.content) < 10000:
                print('(Error 1) Could not connect to Master Control user/password may be incorrect or download link is out of date!')
                return 1;
            else:
                output = open('template.xlsx','wb')
                output.write(r.content)
                output.close()
            return 0;

    #get volume file from SSRS. This uses a package called sspyrs, found at https://pypi.python.org/pypi/sspyrs/0.1a7
    def getVolumeInfo(self):

        #the barcode is simply appended to the end of the URL
        url = 'http://ssrsreports.idtdna.com/REPORTServer/Pages/ReportViewer.aspx?%2fManufacturing%2fSan+Diego%2fPlate+Volume+Information+by+Barcode+ID&rs:Command=Render&BarcodeID='
        url += self.barcode
        report = sspyrs.report(url,self.user,self.password)
        try:
            report.directdown('volumeInfo','EXCEL')
        except ValueError:
            print('\n\n(Error 3) Failed to download volume file. Check that barcode is correct.\n')
            return 1;
        return 0;

    #copies data from volData to template
    def copyData(self):
    
        info = xlrd.open_workbook('volumeInfo.xls')
        template = load_workbook('template.xlsx')

        start = template.active
        start.cell(row=1,column=2).value = self.barcode
        start.cell(row=1,column=9).value = self.user

        #this verifies that the data file is good by checking for a non-blank barcode value in the Cell A1.
        source = info.sheet_by_index(0)
        checkString = source.cell(0,0).value
        delimitedString = checkString.split(' ')
        if(delimitedString[8] == ''):
            print('Error 2: Invalid barcode')
            return 1;
        else:
            dest = template['Raw Data']

            for row in range(2,source.nrows):
                for col in range(0,6):
                    dest.cell(column=col+1, row=(row-1)).value = source.cell(row,col).value

            self.output1 = self.customer + '_' + self.barcode + '.xlsx'
            template.save(filename = self.output1)
            template.close()
        return 0;

    def allOperationsCompletedSuccessfully(self):
        if(self.getTemplate() == 1):
            return False;
        if(self.getVolumeInfo() == 1):
            return False;
        if(self.copyData() == 1):
            return False;
        else:
            return True;
       

    def run(self):
        self.init()
        if(self.allOperationsCompletedSuccessfully()):
            file = self.output1
            os.startfile(file)
            print('\n\n<!!!----No Errors----!!!>\n\n')
            print('Logged in as ' + self.user)
            return 0;
        else:
            return 1;
    

    def start(self):
        self.login()
        self.run()
        #shutdown()
        os.remove('volumeInfo.xls')
        os.remove('template.xlsx')





