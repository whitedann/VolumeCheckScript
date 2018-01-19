import requests
import os
import sspyrs
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
import getpass

#TODO: update input commands..
user = input("user: ")
password = getpass.getpass("pass: ")
barcode = input("barcode: ")
customer = input("Customer: ")

#Work directory
#os.chdir('W:\\Employees\Danny\dev\python')

#Home directory
os.chdir('C:\\Users\Dan\Desktop\idtdocs')

output = customer + '_' + barcode + '.xlsx'

#get template file
payload = {'username':user,'username2':'','initialRequest':'','password':password}

with requests.Session() as s:
    p = s.post('https://idtdna.mastercontrol.com/mc/login/index.cfm?action=login',data=payload)
    r = s.get('https://idtdna.mastercontrol.com/mc/Main/MASTERControl/Organizer/view_file.cfm?id=XBCYSHLULRA4XBIBB7')
    output = open('template.xlsx','wb')
    output.write(r.content)
    output.close()

#get volume file
url = 'http://ssrsreports.idtdna.com/REPORTServer/Pages/ReportViewer.aspx?%2fManufacturing%2fSan+Diego%2fPlate+Volume+Information+by+Barcode+ID&rs:Command=Render&BarcodeID='
url += barcode
report = sspyrs.report(url,user,password)
report.directdown('volumeInfo','Excel')

#copy data (!!change filename!!)
info = load_workbook('volumeInfo.xlsx')
template = load_workbook('template.xlsx')

start = template.active
start.cell['B1'].value = barcode
start.cell['I1'].value = user
start.close()

source = info.worksheets[0]
dest = template.get_sheet_by_name('Raw Data')

for row in range(3,99):
    for col in range(1,7):
        dest.cell(column= col, row = (row-2)).value = source.cell(column=col,row=row).value


template.save(filename = output)
info.close()
template.close()

#TODO: delete template, volumeinfo
os.remove('template.xlsx')
os.remove('volumeInfo.xlsx')

#open prepared file (needs to close?)
file = output
os.startfile(file)



#run vc?


