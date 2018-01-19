
import os
import openpyxl
import cmd
from openpyxl import Workbook
from openpyxl import load_workbook
import cmd
from openpyxl.compat import range
import getpass
os.chdir('C:\\Users\Dan\Desktop\idtdocs')

user = input("User: ")
password = getpass.getpass("Password: ")
barcode = input("Barcode: ")
endUser = input("endUser: ")

output = endUser + '_' + barcode + '.xlsx'

info = load_workbook('Plate Volume Information by Barcode ID - Test.xlsx')
template = load_workbook('testv2.xlsx')

start = template.active
start['B1'].value = barcode
start['I1'].value = user

source = info.worksheets[0]
dest = template.get_sheet_by_name('Raw Data')

for row in range(3,99):
    for col in range(1,7):
        dest.cell(column= col, row = (row-2)).value = source.cell(column=col,row=row).value

template.save(filename = output)
info.close()
template.close()

file = output
os.startfile(output)

##Todo
#def validateInput(str):
 #   if(str == barcodeIn):
  #      buffer = input('Barcode ID: ')
   #     if(buffer.length == 7):
    #        barcode = buffer
     #       break
      #  else:
       #     print('invalid barcode')
            
    



                                                            
