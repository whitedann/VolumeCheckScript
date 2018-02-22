import sys
sys.stdout.write('Loading dependencies..')

import requests
import os
import sspyrs
import openpyxl
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
import getpass
print('done!')

def init():
    global user
    global password
    global barcode
    global customer
    user = input('type user (don\'t scan badge): ')
    password = getpass.getpass('type password (hidden): ')
    barcode = input('Enter BARCODE: ')
    customer = input('Enter Customer (optional): ')
    
    #Work directory
    os.chdir('W:\\Employees\Danny\dev')

    #Home directory
    #os.chdir('C:\\Users\Dan\Desktop\idtdocs')
    return 0;

#get template file
def getTemplate():
    payload = {'username':user,'username2':'','initialRequest':'','password':password}

    with requests.Session() as s:
        p = s.post('https://idtdna.mastercontrol.com/mc/login/index.cfm?action=login',data=payload)
        r = s.get('https://idtdna.mastercontrol.com/mc/Main/MASTERControl/Organizer/view_file.cfm?id=XBCYSHLULRA4XBIBB7')
        if len(r.content) < 10000:
            print('Could not connect to Master Control user/password may be incorrect')
            start()
        else:
            output = open('template.xlsx','wb')
            output.write(r.content)
            output.close()
        return 0;

#get volume file
def getVolumeInfo():
    url = 'http://ssrsreports.idtdna.com/REPORTServer/Pages/ReportViewer.aspx?%2fManufacturing%2fSan+Diego%2fPlate+Volume+Information+by+Barcode+ID&rs:Command=Render&BarcodeID='
    url += barcode
    report = sspyrs.report(url,user,password)
    try:
        report.directdown('volumeInfo','EXCEL')
    except ValueError:
        print('\n\nFailed to download volume file. Check that barcode is correct.\n')
        start()
    return 0;

#copy data
def copyData():
    global output1
    info = xlrd.open_workbook('volumeInfo.xls')
    template = load_workbook('template.xlsx')

    start = template.active
    start.cell(row=1,column=2).value = barcode
    start.cell(row=1,column=9).value = user

    source = info.sheet_by_index(0)
    checkString = source.cell(0,0).value
    delimitedString = checkString.split(' ')
    if(delimitedString[8] == ''):
        print('invalid barcode')
        return 1;
    else:
        dest = template['Raw Data']

        for row in range(2,source.nrows):
            for col in range(0,6):
                dest.cell(column=col+1, row=(row-1)).value = source.cell(row,col).value

        output1 = customer + '_' + barcode + '.xlsx'
        template.save(filename = output1)
        template.close()
    return 0;

def start():

    init()
    getTemplate()
    getVolumeInfo()

    if(copyData() == 0):
        file = output1
        os.startfile(file)
        start()
        print('Success')
    else:
        start()
    os.remove('volumeInfo.xls')
    os.remove('template.xlsx')
    return 0;


start()


